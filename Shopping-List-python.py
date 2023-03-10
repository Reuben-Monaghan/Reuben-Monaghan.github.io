# Imports
import openpyxl
import os
import pyinputplus as pyip
import tkinter as tk
from tkinter import ttk
from collections import Counter
import datetime
global file_path 
file_path = ''

# Open the recipe data excel file

def open_Recipes_xl():
    os.chdir(file_path)

    workbook = openpyxl.load_workbook('Recipes.xlsx')
    sheet1 = workbook['Sheet1']
    return sheet1,workbook

# Create recipe class
class recipe:
    
    # instance attributes, takes recipe name the serving size of the recipe and the ingredients dictionary {ingredient:quantity}
    
    def __init__(self, name, servings, ingredients):
        self.name = name
        self.servings = servings        
        self.ingredients = ingredients
        
        
    
    
    def printname(self):
        print(self.name)
        
    def add_recipe(self):
        
        sheet1,workbook = open_Recipes_xl()
        
        recipe_row = len(recipe_list())+1
        
        sheet1.cell(row=recipe_row, column=1).value = self.name
        
        sheet1.cell(row=recipe_row, column=2).value = int(self.servings)
        
        sheet1.cell(row=recipe_row, column=3).value = len(self.ingredients)
        
        for i in range(len(self.ingredients)):
            
    
            sheet1.cell(row=recipe_row, column=((i*3)+4)).value = self.ingredients[i][0]
            sheet1.cell(row=recipe_row, column=((i*3)+5)).value = self.ingredients[i][1]
            sheet1.cell(row=recipe_row, column=((i*3)+6)).value = self.ingredients[i][2]
        workbook.save('Recipes.xlsx')
        
        
    def add_ingredients(self):

        sheet1,workbook = open_Recipes_xl()

        recipe_row = recipe_list().index(r.name) + 1
        ingredient_range = get_ingredient_range(recipe_row)
        new_ingredient_range = (ingredient_range[1]+1),(ingredient_range[1]+4)

        while True:

            ing  = pyip.inputStr(prompt='Enter ingredient name: ')
            quant  = pyip.inputNum(prompt='Enter ingredient quantity: ')
            measure  = pyip.inputStr(prompt='Enter ingredient measurement: ')
            check = pyip.inputStr(prompt=f'{ing}-{quant}{measure}, is this correct? Y/N: ')

            if check.upper() == 'Y':
                break
            else:
                continue

          # make a list of ing,quant,measure then add to new_ingredient_range.value in excel sheet.

        add_list=[ing, quant, measure]

        for i in range(new_ingredient_range[0],new_ingredient_range[1]):
            sheet1.cell(row=recipe_row, column=i).value = add_list[i-new_ingredient_range[0]]

        workbook.save('Recipes.xlsx')

    def delete_ingredients(self):
        sheet1,workbook = open_Recipes_xl()

        # getting all the info needed
        recipe_row = recipe_list().index(r.name) + 1
        recipe_info = get_recipe_info(recipe_row)
        ingredient_range = get_ingredient_range(recipe_row)

        ingredient_names_list = recipe_info[-1][::3]
        ingredient_list = recipe_info[-1]


        for i in range(3):
            print(f'ingredients in {r.name}: {ingredient_names_list}')
            del_ing  = pyip.inputStr(prompt='Enter ingredient name to delete: ')

            # going to need to add a way of keeping all strings uniform i.e. capitalise all ingredients

            if del_ing in ingredient_names_list:
                break
            else:
                print(f'{del_ing} is not in {r.name}')
                continue


        if del_ing in ingredient_names_list:
            check = pyip.inputStr(prompt=f'are you sure you want to delete {del_ing} from {r.name} ingredients? Y/N: ')

            #if yes, delete ingredients.
            if check.upper() == 'Y':

                #delete ingredient
                columns_to_delete = (ingredient_list.index(del_ing)+4),(ingredient_list.index(del_ing)+7)        
                DeleteCellsxl(recipe_row,columns_to_delete)

                #reduce the number of ingredients by one.
                sheet1.cell(row=recipe_row, column=3).value -=1     

                #copy remaining igredients - find the last ingredient cell and the last cell of the deleted ingredient.
                # all cells in that range should be move to the left by three (-3) then delete the excess cells

                columns_to_copy = (columns_to_delete[1]),(len(ingredient_list)+4)
                columns_to_paste = (columns_to_delete[0]),(columns_to_copy[1]-3)
                CopyandPastexl(recipe_row,columns_to_copy,columns_to_paste)

                #delete excess cells
                columns_to_delete = (columns_to_paste[1]),(columns_to_copy[1])
                DeleteCellsxl(recipe_row,columns_to_delete)


                print(f'{del_ing}, deleted from: {r.name}')

        else:
            print('You did not select an existing ingredient to delete')

        workbook.save('Recipes.xlsx')

# Open the excel file and get a list of the recipe names
def recipe_list():
    recipes_list=[]
    
    sheet1, workbook = open_Recipes_xl()
    for row in sheet1.values:
        if row[0] != None:
            recipes_list.append(row[0])
        else:
            continue
    
    return recipes_list      

def get_ingredients(selected_recipes):
    
    #This uses the recipe_list() and gets the ingredients for the selected recipes using get_recipe_info() to return a list
    # of ingredients. -- this needs to be grouped using join_ingredients() before the duplicates can be 
    # merged using merge_duplicates()
    
    r = recipe_list()
    ingredients = []
    for i in selected_recipes:
        for j in get_recipe_info(r.index(i)+1)[-1]:
            
            ingredients.append(j)
        
    return ingredients

def join_ingredients(ingredients):
    
    # This groups the ingredients into [ingredient, quantity, measurement] and returns a list of grouped items
    
    # Initialize the result list
    result = []

    # Iterate over the list in groups of three
    for i in range(0, len(ingredients), 3):
        
        # Create a range object that starts at i and ends at i + 3
        group = ingredients[i:i+3]
        
        # Concatenate the items in the group using the join function
        concatenated = " ".join(group)
        
        # Append the resulting string to the result list
        result.append(group)

    return result

def merge_duplicates(grouped_ingredients):
    
    #create a list for the merged duplicates
    merged_duplicates = []
    
    #create a counter object that counts the number of instances of each ingredient in grouped_ingredients
    c = Counter(i[0] for i in grouped_ingredients)
    
    #create a list of any duplicate ingredients in c
    duplicates = []
    for i in c:
        if (c[i]) > 1:
            duplicates.append([i][0])
            
    for i in duplicates:
    #for each duplicate...
    
        duplicates_to_merge = []
        


        for j in grouped_ingredients:

            #for each item in grouped_ingredients check if it is a duplicate        
            if j[0] == i:

                #append duplicates to a list to be merged
                duplicates_to_merge.append(j)

        #get the quantity values of the duplicates
        quantities_of_duplicate = [x[1] for x in duplicates_to_merge]

        # get the sum() of the duplicate values
        result = sum([int(i) for i in quantities_of_duplicate])

        #create a list with the merged duplicate with the sum() as the quantity
        merged_duplicates.append([duplicates_to_merge[0][0],str(result),duplicates_to_merge[0][2]])
                
    # create a new_ingredient_list:
    # first remove duplicates from grouped_ingredients 
    
    new_ingredient_list = [x for x in grouped_ingredients if x[0] not in duplicates]
    # then append the merged ingredients to the newlist
    for i in merged_duplicates:
        new_ingredient_list.append(i)
    # return the new_ingredient_list with original duplicates removed and merged duplicates added
    return new_ingredient_list

def write_shopping_list(ingredient_list):
    os.chdir(file_path)
    now = datetime.datetime.now()
    date_string = now.strftime('%d-%m-%Y')
    date_string
    # Open the file in write mode
    with open(f'Shopping list - {date_string}.txt', 'w') as f:
        # Write each item to the file as a line
        for item in ingredient_list:
            f.writelines(' '.join(item) + '\n')

    # The file is automatically closed when the block ends

def DeleteCellsxl(row,del_range):
    # this changes the value of a range of cells to None
    
    for cell in range(del_range[0],del_range[1]):
        sheet1.cell(row=row, column=cell).value = None
            
def CopyandPastexl(row,copy_range,paste_range):
    # This copies a range of cells (copy_range) within a row (row), then pastes the values into a new range (paste_range)
        
    copy_list=[]
    
    for cell in range(copy_range[0],copy_range[1]+1):
        copy_list.append(sheet1.cell(row=row, column=cell).value)
    
    for cell in range(paste_range[0],paste_range[1]+1)[::-1]:
        sheet1.cell(row=row, column=cell).value = copy_list.pop()

def submit(recipes,checkboxes):
    # A function to be called when the "Submit" button is clicked
    # Create a new list to store the selected recipes
    
    
    selected = []
    for i, recipe in enumerate(recipes):
        
                
        if checkboxes[i].get() == 1:
            selected.append(recipe)
    
        # Assign the selected recipes to the global variable
    global selected_recipes
    selected_recipes = selected
    return selected_recipes
    
def Create_Shopping_List(menu_window):
    #Called when the "Create shopping list" button is clicked from the main menu
    
    # Close the menu window
    menu_window.destroy()
    
    # Create Select recipes window
    select_recipes_window = tk.Tk()
    
    # Create a list of recipes
    # Open the excel document (open_Recipes_xl()) <- this is in the recipe_list() function. and create a recipe list using
    # recipe_list()
    
    recipes = recipe_list()
    
    # Create a list of tkinter variables for the checkboxes
    checkboxes = []
    for recipe in recipes:
        var = tk.IntVar()
        checkboxes.append(var)
    
    # Create the checkboxes and "Submit" button
    for i, recipe in enumerate(recipes):
        tk.Checkbutton(select_recipes_window, text=recipe, variable=checkboxes[i]).pack()

    # Create a "Submit" button
    submit_button = tk.Button(select_recipes_window, text='Submit', command=lambda:[submit(recipes,checkboxes),select_recipes_window.destroy(),Main_Menu()])
    submit_button.pack()


    # Run the main loop
    select_recipes_window.mainloop()
    
    # trying to string all the functions together to create a shopping list
    # selected_recipes >> A - get_ingredients() >> B - join_ingredients >> C - merge_duplicates >> D - write_shopping_list
    sheet1,workbook = open_Recipes_xl()
    A = get_ingredients(selected_recipes)
    B = join_ingredients(A)
    C = merge_duplicates(B)
    write_shopping_list(C)

def get_ingredient_range(selection_row):
    sheet1,workbook = open_Recipes_xl()
    # find the number of ingredients, * by 3 to get ingredient,quantity,measurement + 4 to select the correct columns
    num_ingredients = sheet1.cell(row=selection_row, column=3).value
    ingredient_range = 4,(3+(num_ingredients*3))
    return ingredient_range

def get_recipe_info(selection_row):
    # add 1 to selection_row to convert from recipes_list.index to excel row number
    sheet1,workbook = open_Recipes_xl()
        
    #display recipe name, serving size, and number of ingredients
    recipe_name,serving_size,num_ingredients = sheet1.cell(row=selection_row, column=1).value,sheet1.cell(row=selection_row, column=2).value,sheet1.cell(row=selection_row, column=3).value
            
    # find the number of ingredients
    ingredient_range = get_ingredient_range(selection_row)
    
    
    ingredients = []
    for i in range(ingredient_range[0],ingredient_range[1]+1):
        ingredients.append(str(sheet1.cell(row=selection_row, column=i).value))
        
    
    return recipe_name,serving_size,num_ingredients,ingredients

def delete_recipes(recipes,selected_recipes):
    
    #this deletes the rows using the index of selected_recipes in the recipes list
    #open the workbook and delete the rows from selected_recipes.
    
    sheet1,workbook = open_Recipes_xl()
    
    for i in selected_recipes:
        
        sheet1.delete_rows((recipes.index(i)+1))
    
    workbook.save('Recipes.xlsx')    
    
def delete_recipes_button(menu_window):
    
    global selected_recipes
    selected_recipes = []
    
    # Close the menu window
    menu_window.destroy()
    
    # Create Select recipes window
    delete_recipes_window = tk.Tk()
    
    # Create a list of recipes
    # Open the excel document (open_Recipes_xl()) <- this is in the recipe_list() function. and create a recipe list using
    # recipe_list()
    recipes=[]
    recipes = recipe_list()
    
    # Create a list of tkinter variables for the checkboxes
    checkboxes = []
    for recipe in recipes:
        var = tk.IntVar()
        checkboxes.append(var)
    
    # Create the checkboxes and "Submit" button
    for i, recipe in enumerate(recipes):
        tk.Checkbutton(delete_recipes_window, text=recipe, variable=checkboxes[i]).pack()

    # Create a "Submit" button
    submit_button = tk.Button(delete_recipes_window, text='Submit', command=lambda:[submit(recipes,checkboxes),delete_recipes(recipes,selected_recipes),delete_recipes_window.destroy(),Main_Menu()])
    submit_button.pack()

    
    # Run the main loop
    delete_recipes_window.mainloop()

def add_recipe(menu_window):
    #Called when the "Add recipe" button is clicked from the main menu
    
    
    global recipe_name
    recipe_name = ''
    global servings
    servings = ''
    global ingredients
    ingredients = []
    global ingredient_num
    ingredient_num = 0
    
    global I
    global Q
    global M
    I = ''
    Q = ''
    M = ''
    
    # Initialize the widget_y variable with the y position of the first set of widgets
    global widget_y
    widget_y = 100
    
    # Close the menu window
    menu_window.destroy()
    
    # Create add recipe window
    add_recipe_window = tk.Tk()
    
     # Create the frame to hold the canvas and the scrollbar
    frame = tk.Frame(add_recipe_window)
    frame.pack()

    
    # Create the canvas widget and the scrollbar
    canvas = tk.Canvas(add_recipe_window, height = 800, width = 770 )
    canvas.pack(side = "left")
    scrollbar = ttk.Scrollbar(add_recipe_window, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas['yscrollcommand'] = scrollbar.set
    
       
    
    # Create a label for the title
    ttk.Label(add_recipe_window, text="Add New Recipe")
    
    #Create recipe name label and entry box
    recipe_name_label = ttk.Label(add_recipe_window, text="Recipe Name:")
    recipe_name_input_box = tk.Entry(add_recipe_window,textvariable=recipe_name )
    
    #Create number of servings label and entry box
    Number_of_servings_label = ttk.Label(add_recipe_window, text="Number of servings:")
    Number_of_servings_input_box = tk.Entry(add_recipe_window)
    
    
    # labels and entry boxes
    Ingredient_label = ttk.Label(add_recipe_window, text="Ingredient:")
    Ingredient_input_box = tk.Entry(add_recipe_window,textvariable=I )
    
    Quantity_label = ttk.Label(add_recipe_window, text="Quantity:")
    Quantity_input_box = tk.Entry(add_recipe_window)
    
    Measurement_label = ttk.Label(add_recipe_window, text="Measurement:")
    Measurement_input_box = tk.Entry(add_recipe_window, textvariable='measurement')

    # locations of labels and entry boxes
    canvas.create_window((50, 30), window=recipe_name_label)
    canvas.create_window((200, 30), window=recipe_name_input_box)
    canvas.create_window((69, 60), window=Number_of_servings_label)
    canvas.create_window((200, 60), window=Number_of_servings_input_box)    
    canvas.create_window((50, 100), window=Ingredient_label)
    canvas.create_window((150, 100), window=Ingredient_input_box)
    canvas.create_window((255, 100), window=Quantity_label)
    canvas.create_window((350, 100), window=Quantity_input_box)
    canvas.create_window((490, 100), window=Measurement_label)
    canvas.create_window((600, 100), window=Measurement_input_box)
    
    
    # Create a button to Add Recipe 
    ttk.Button(add_recipe_window, text="Add Recipe", command=lambda:[add_recipe_button_command(),
                                                                     add_recipe_window.destroy(),
                                                                    Main_Menu()]).place(x=650, y=750)
    
    # Create a button to cancel
    ttk.Button(add_recipe_window, text="Cancel", command=lambda:[add_recipe_window.destroy(),Main_Menu()]).place(x=50, y=750)
    
    
    
    
    # This function creates new widgets and adds them below the previous ones (this is so i can add multiple new ingredients)
    def add_ingredient_widgets(I,Q,M):
        #global added_ingredient_label
        global widget_y
        global ingredient_num

        widget_y += 30
        ingredient_num +=1

        ingredient_num_label = ttk.Label(add_recipe_window, text=f"{ingredient_num}:")
        
        added_ingredient_label = ttk.Label(add_recipe_window, text=f"Ingredient: {I}\t  Quantity: {Q} \t Measurement: {M}")
        
        # Create the windows inside the canvas and position them using widget_y
        canvas.create_window((160, widget_y), window=ingredient_num_label)
        canvas.create_window((160, widget_y), window=added_ingredient_label)
                
        

    add_ingredient_button = ttk.Button(add_recipe_window, text="Add Ingredient", command=lambda:[add_ingredient_button_command()
                                                                                                 ,add_ingredient_widgets(I,Q,M)])

    # Create the button inside the canvas and position it
    canvas.create_window((725, 100), window=add_ingredient_button)
    


    def add_ingredient_button_command():
        
        # This will use .get() to take the information in the add igredient entry boxes and store it in a list
        global I
        global Q
        global M
        global ingredients
        
        I = Ingredient_input_box.get()
        Q = Quantity_input_box.get()
        M = Measurement_input_box.get()
        
        ingredients.append([I,Q,M])
        
        # this deletes the contents of the entry boxes
        Ingredient_input_box.delete(0, tk.END)
        Quantity_input_box.delete(0, tk.END)
        Measurement_input_box.delete(0, tk.END)
        
        return I,Q,M
    
    def add_recipe_button_command():
        global recipe_name
        global servings
        
        recipe_name = recipe_name_input_box.get()
        servings = Number_of_servings_input_box.get()
        
        m = recipe(recipe_name,servings,ingredients) 
        m.add_recipe()

def Main_Menu():
    # Tkinter Main menu
    # This opens the main menu then calls functions by clicking buttons.

    # Create a list to store the selected recipes

    selected_recipes = []

    #open excel file
    sheet1,workbook = open_Recipes_xl()

    # Create the main window
    menu_window = tk.Tk()

    # Create the frame/format
    frame = ttk.Frame(menu_window, padding=10)
    frame.grid()

    # Create a label for the title
    ttk.Label(frame, text="Shopping List Menu").grid(column=1, row=0)

    # Create a button to add a Add Recipe 
    ttk.Button(frame, text="Add Recipe", command=lambda:[add_recipe(menu_window)]).grid(column=0, row=1)

    # Create a button to Edit Recipe
    ttk.Button(frame, text="Edit Recipe", command=menu_window.destroy).grid(column=0, row=2)

    # Create a button to Delete Recipe
    ttk.Button(frame, text="Delete Recipe", command=lambda:[delete_recipes_button(menu_window)]).grid(column=0, row=3)

    # Create a button to Create Shopping List
    ttk.Button(frame, text="Create Shopping List", command=lambda:[Create_Shopping_List(menu_window)]).grid(column=3, row=1)

    # Create a button to quit
    ttk.Button(frame, text="Quit", command=menu_window.destroy).grid(column=1, row=4)

    menu_window.mainloop()

# Run the Main_Menu Function to start the program

Main_Menu()